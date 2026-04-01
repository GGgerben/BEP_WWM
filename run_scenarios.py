
# ------------------------------------------------------------
#
# Runs 27 scenarios:
# Wealth (Arm/Gemiddeld/Rijk) x Experience (Nooit/Een keer/Vaker dan een keer) x N (10/100/1000)
#
# Output (Excel):
# 1) ScenarioSummary  
# 2) MeasureTop5      
#
# ------------------------------------------------------------


import random
from collections import Counter
from statistics import mean, stdev
from typing import Dict, List, Tuple

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from classes.measures import measures  # lijst met maatregelen
from classes.hazard_generator import floods
from classes.scenario_initialisation import initialise_scenario_population

from data.houses_dict import houses_dict, generate_houses_from_agents  # type: ignore



# Wat wil ik weten?

def unique_measures_of_agent(agent) -> set[str]:
    """Unieke maatregelen minimaal een keer geadopteerd, dus hoeveel soorten."""
    return set(m.name for (m, r) in getattr(agent, "adopted_measures", []))


def total_purchases_of_agent(agent) -> int:
    """Totaal aantal aankppen inclusief herhalingen."""
    return len(getattr(agent, "adopted_measures", []))


def adoption_rates(agents) -> Dict[str, float]:
    """ Per maatregel, het aantal keer geadopteerd.
    """
    N = len(agents)
    c = Counter()
    for a in agents:
        for name in unique_measures_of_agent(a):
            c[name] += 1
    return {name: c[name] / N for name in c}


def scenario_core_stats(agents) -> Dict[str, float]: # Geeft gemiddeld aantal maatregelen per agent en gemiddeld aantal unieke maatregelen per agent, en satis.
    
    unique_counts = [len(unique_measures_of_agent(a)) for a in agents]
    total_counts = [total_purchases_of_agent(a) for a in agents]
    sats = [float(getattr(a, "satisfaction", 0.0)) for a in agents]

    return {
        "mean_unique_measures_per_agent": mean(unique_counts),
        "mean_total_purchases_per_agent": mean(total_counts),
        "mean_satisfaction": mean(sats),
    }

def purchase_counts(agents) -> Counter:  # Geeft per maatregel aan hoe vaak deze is gekocht, inclusief herhalingen
    
    c = Counter()
    for a in agents:
        for (m, r) in getattr(a, "adopted_measures", []):
            c[m.name] += 1
    return c

def run_one_simulation(
    wealth_class: str,
    experience_level: str,
    n_agents: int,
    seed: int,
    n_rounds: int = 4,
):
    """
    Runt model voor 1 herhaling en geeft agenten na n_rounds.
    """
    random.seed(seed)

    # 1) Scenario specifieke populatie
    agents = initialise_scenario_population(
        n=n_agents,
        seed=seed,
        wealth_class=wealth_class,
        experience_level=experience_level,
    )

    # 2) Huizenmarkt, want volledig onafhankelijke simulatie
    big_houses_dict = generate_houses_from_agents(
        houses_dict,
        agents,
        target_n_houses=2000,
        seed=seed,
        affordability_quantile=0.95,
        house_price_quantile=0.20,
        jitter=0.10,
    )

    # 3) Run rounds
    for round_nr in range(1, n_rounds + 1):
        flood_results = floods()
        for agent in agents:
            agent.step(big_houses_dict, measures, flood_results, current_round=round_nr)

    return agents


# Scenario experiment (27 scenarios)
def run_all_scenarios(
    n_reps: int = 10,
    base_seed: int = 1000,
    n_rounds: int = 4,
    top_k_measures: int = 5,
) -> Tuple[List[dict], List[dict]]:
    wealth_levels = ["Rijk", "Gemiddeld", "Arm"]
    exp_levels = ["Nooit", "Een keer", "Vaker dan een keer"]
    Ns = [10, 100, 1000]

    scenario_rows: List[dict] = []
    topk_rows: List[dict] = []

    scenario_id = 0

    for w in wealth_levels:
        for e in exp_levels:
            for N in Ns:
                scenario_id += 1

                # Lege lijsten voor het opslaan van de gevonden data
                rep_unique = []
                rep_total = []
                rep_sat = []
                rep_rates_list: List[Dict[str, float]] = []
                rep_purchase_counts_list: List[Counter] = [] #toegevoegd

                for rep in range(n_reps):
                    seed = base_seed + scenario_id * 10_000 + rep
                    agents = run_one_simulation(w, e, N, seed, n_rounds=n_rounds)

                    stats = scenario_core_stats(agents)
                    rep_unique.append(stats["mean_unique_measures_per_agent"])
                    rep_total.append(stats["mean_total_purchases_per_agent"])
                    rep_sat.append(stats["mean_satisfaction"])

                    rep_rates_list.append(adoption_rates(agents))
                    rep_purchase_counts_list.append(purchase_counts(agents)) #toegevoegd

                # resultaten van alle runs samen nemen tot gemiddelden
                def _avg(x): return float(mean(x))
                def _sd(x): return float(stdev(x)) if len(x) > 1 else 0.0

                scenario_rows.append({
                    "scenario": scenario_id,
                    "wealth": w,
                    "experience": e,
                    "N": N,
                    "rounds": n_rounds,
                    "reps": n_reps,
                    "avg_unique_measures_per_agent": _avg(rep_unique),
                    "sd_unique_measures_per_agent": _sd(rep_unique),
                    "avg_total_purchases_per_agent": _avg(rep_total),
                    "sd_total_purchases_per_agent": _sd(rep_total),
                    "avg_satisfaction": _avg(rep_sat),
                    "sd_satisfaction": _sd(rep_sat),
                })

                # Top meest gekozen maatregelen (op basis van hoe vaak gekozen per agent)
                all_measures = set()
                for d in rep_purchase_counts_list:
                    all_measures |= set(d.keys())

                measure_intensities = []
                for m in all_measures:
                    # per herhaling: (totaal aankopen van m) / N  => gemiddeld aantal aankopen per agent
                    avg_purchases_per_agent = float(
                        mean(d.get(m, 0) / N for d in rep_purchase_counts_list)
                    )
                    measure_intensities.append((m, avg_purchases_per_agent))

                measure_intensities.sort(key=lambda x: x[1], reverse=True) #sorteert maatregelen op gemiddelde aankoopintensiteit per agent

                for rank, (m, avg_purchases_per_agent) in enumerate(measure_intensities[:top_k_measures], start=1):
                    topk_rows.append({
                        "scenario": scenario_id,
                        "wealth": w,
                        "experience": e,
                        "N": N,
                        "rank": rank,
                        "measure": m,
                        "avg_purchases_per_agent": avg_purchases_per_agent,
                    })
                

                print(f"Done scenario {scenario_id:02d} | {w:9s} | {e:18s} | N={N} | reps={n_reps}")

    return scenario_rows, topk_rows


# Excel export (2 sheets)

def autofit_columns(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                v = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(v))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 45)


def write_table(ws, rows: List[dict], table_name: str, style_name: str = "TableStyleMedium2"):
    if not rows:
        raise ValueError("No rows to write.")

    headers = list(rows[0].keys())
    ws.append(headers)
    for r in rows:
        ws.append([r.get(h) for h in headers])

    end_row = ws.max_row
    end_col = ws.max_column
    ref = f"A1:{get_column_letter(end_col)}{end_row}"

    tab = Table(displayName=table_name, ref=ref)
    tab.tableStyleInfo = TableStyleInfo(
        name=style_name,
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tab)
    autofit_columns(ws)


def export_excel(path: str, scenario_rows: List[dict], topk_rows: List[dict]):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "ScenarioSummary"

    write_table(ws1, scenario_rows, table_name="ScenarioSummaryTable")

    ws2 = wb.create_sheet("MeasureTopK")
    write_table(ws2, topk_rows, table_name="MeasureTopKTable")

    wb.save(path)


if __name__ == "__main__":
    N_REPS = 100
    BASE_SEED = 1000
    N_ROUNDS = 4
    TOP_K = 5

    scenario_rows, topk_rows = run_all_scenarios(
        n_reps=N_REPS,
        base_seed=BASE_SEED,
        n_rounds=N_ROUNDS,
        top_k_measures=TOP_K,
    )

    out_file = "ScenarioResults.xlsx"
    export_excel(out_file, scenario_rows, topk_rows)

    print("\nSaved:")
    print(f"- {out_file}")
    print("Sheets: ScenarioSummary, MeasureTopK")

